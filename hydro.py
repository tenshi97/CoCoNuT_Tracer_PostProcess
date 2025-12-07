import numpy as np

import h5reader
import matplotlib.pyplot as plt
import math
import re
import copy
from collections import deque
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

queue = deque()
pending_queue = deque()
h=h5reader.hydrof(model='s16')
f = open("test.txt","w")
g = open("output.txt","w")

wb = Workbook()

xif = h.xzr()
yif = h.yzr()
def grid_finder(x,y):
    st = 0
    ed = r_len - 1
    x_cell = -1
    y_cell = -1
    while(st<ed):
        mid = (st+ed)//2
        if(x<xif[mid]):
            ed = mid
        else:
            st = mid+1
    x_cell = st
    st = 0
    ed = theta_len - 1
    while(st<ed):
        mid = (st+ed)//2
        if(y<yif[mid]):
            ed = mid
        else:
            st = mid+1
    y_cell = st
    return (x_cell,y_cell)

t = h.ngroups-1

h.set_index(t)
#for t in range(0,y.ngroups):
#    y.set_index(t)
#    tracer_ids = np.array([int(round(x)) for x in y.tracer_id()])
#    for i in range(len(tracer_ids)):
#        gid = tracer_ids[i]
#        tracer_data[gid].append((y.time(),y.tracer_x()[i],y.tracer_y()[i]))
r_len = len(h.xzn())
theta_len = len(h.yzn())
empty_dict = {}
array_2D = [[] for x in range(r_len)]
for row in array_2D:
    for i in range(theta_len):
        row.append({})

#Init
tracer_ids = np.array([int(round(x)) for x in h.tracer_id()])
tracer_num = len(tracer_ids)
tracer_data = []
for i in range(tracer_num):
    (x_cell,y_cell) = grid_finder(h.tracer_x()[i],h.tracer_y()[i])
    tracer_data.append((tracer_ids[i],h.tracer_x()[i],h.tracer_y()[i],x_cell,y_cell))

for i in range(tracer_num):
    tracer = tracer_data[i]
    id = tracer[0]
    x = tracer[3]
    y = tracer[4]
    grid = array_2D[x][y]
    grid[id] = 0
    num = len(grid)
    for key in grid.keys():
        grid[key] = 1/num

for x in range(r_len):
    for y in range(theta_len):
        if(len(array_2D[x][y])!=0):
            cell = ws.cell(row=x + 1, column=y + 1)
            color = f"{153:02X}{217:02X}{234:02X}"
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.fill = fill
            cv = ""
            grid = array_2D[x][y]
            for key in grid:
                cv += f"{int(key)}:{grid[key]:.3f}\n"
            cell.value = cv
            queue.append((x,y))
print(queue)
#Init
def get_neighbour_grid(x,y):
    result = []
    if(x-1>=0):
        result.append((x-1,y))
    if(y-1>=0):
        result.append((x,y-1))
    if(x+1<r_len):
        result.append((x+1,y))
    if(y+1<theta_len):
        result.append((x,y+1))
    return result
step = 0
f.write(f"{queue}\n")
while(True):
    if(len(queue)==0):
        if(len(pending_queue)==0):
            break
        else:
            for grid in pending_queue:
                x = grid[0]
                y = grid[1]
                total = 0
                for value in array_2D[x][y].values():
                    total += value
                for key in array_2D[x][y].keys():
                    array_2D[x][y][key]/=total
            queue.extend(pending_queue)
            #f.write(f"{queue}\n")
            pending_queue.clear()
    step += 1
    grid_pop = queue.popleft()
    #f.write(f"-----Step{step}----\n")
    #f.write(f"First Grid:{grid_pop}\n")
    xp = grid_pop[0]
    yp = grid_pop[1]
    neighbours = get_neighbour_grid(grid_pop[0],grid_pop[1])
    for grid in neighbours:
        x = grid[0]
        y = grid[1]
        if(len(array_2D[x][y])!=0):
            if(grid in pending_queue):
                #pending_queue.append(grid)
                array_2D[x][y].update(array_2D[xp][yp])
        else:
            if(not grid in pending_queue):
                pending_queue.append(grid)
                array_2D[x][y].update(array_2D[xp][yp])
Mass_Assignment = [0.0 for x in range(tracer_num)]
c = 29979245800
for x in range(r_len):
    for y in range(theta_len):
        grid = array_2D[x][y]
        cell = ws.cell(row=x+1,column=y+1)
        cv = ""
        for key in grid:
            cv+=f"{int(key)}:{grid[key]:.3f}\n"
        if cell.value is None or cell.value == "":
            cell.value = cv
        cell.alignment = Alignment(wrapText=True)

        r_1 = h.xzl()[x]
        r_2 = h.xzr()[x]
        theta_1 = h.yzl()[y]
        theta_2 = h.yzr()[y]
        theta_mid = h.yzn()[y]
        DV = 2*np.pi*(np.cos(theta_1)-np.cos(theta_2))*(r_2**3-r_1**3)/3
        LF = 1.0/np.sqrt(1-(h.vex()[x][y][0]**2+h.vey()[x][y][0]**2)/(c**2))    #Lawrence Factor
        DM = h.den()[x][y][0]*DV*h.phi()[x][y][0]*LF

        for key in grid:
            t_id = int(key)
            t_weight = grid[key]
            Mass_Assignment[t_id-1] += t_weight*DM
wb.save("output.xlsx")
Mass_Sun = 1.98855e33   # UNIT:grams
Total_Mass = 0.0
for i in range(tracer_num):
    g.write(f"{i+1}: {Mass_Assignment[i]}\n")
    Total_Mass += Mass_Assignment[i]
MS = Total_Mass/Mass_Sun
tracer_alltime_data = [[] for x in range(tracer_num)]
for t in range(h.ngroups):
    h.set_index(t)






g.write(f"Total Mass:{Total_Mass}({MS:.3f}of Solar Mass)")





# data = []
# procs = 48
# tracer_per_proc = 40
# total_tracer = procs*tracer_per_proc
# tracer_data = [[] for x in range(0,total_tracer)]
# time_data = [[] for x in range(0,3700)]
# for i in range(0,48):
#     print(i)
#     timestep_p = 0
#     with open("F:/compastro/hdf5output/debug/filen"+str(i),mode='r') as file:
#         data_proc = []
#         tracer_num_record = []
#         current_timestep = 0
#         data_proc_timestep = {"timestep":0,"tracers": []}
#         local_id = 0
#         global_id = 0
#         pos_r = 0
#         cellindex_r = 0
#         pos_theta = 0
#         cellindex_theta = 0
#         for line in file:
#             if(line.find("Cell")!=-1):
#                 continue
#             if(line.find("IPROC")!=-1):
#                 continue
#             if(line.find("v_r")!=-1 or line.find("v_theta")!=-1):
#                 continue
#             if(line.find("INIT")!=-1):
#                 continue
#             pattern1 = re.compile(r"TRACER NO =\s+(\d+)\s+TIME =\s+(\d+.\d+E?[-+]?\d+)\s+(\d+)")
#             match1 = pattern1.search(line)
#             if(match1):
#                 local_id = int(match1.group(1))
#                 time = float(match1.group(2))
#                 timestep = int(match1.group(3))
#                 if(current_timestep!=timestep):
#                     if(current_timestep!=0):
#                         timestep_p+=1
#                         tracer_num_record.append(len(data_proc_timestep["tracers"]))
#                         data_proc.append(copy.deepcopy(data_proc_timestep))
#                     data_proc_timestep["timestep"] = timestep
#                     data_proc_timestep["time"] = time
#                     data_proc_timestep["tracers"] = []
#                     current_timestep = timestep
#                 continue
#             pattern2 = re.compile(r"GLOBAL ID=\s+(\d+)")
#             match2 = pattern2.search(line)
#             if match2:
#                 global_id = int(match2.group(1))
#                 continue
#             pattern3 = re.compile(r"position r = \s+(\d+.\d+E?-?\d+)\s+(\d+)")
#             match3 = pattern3.search(line)
#             if match3:
#                 pos_r = float(match3.group(1))
#                 cellindex_r = int(match3.group(2))
#                 continue
#             pattern4 = re.compile(r"position theta = \s+(\d+.\d+E?-?\d+)\s+(\d+)")
#             match4 =  pattern4.search(line)
#             if match4:
#                 pos_theta = float(match4.group(1))
#                 cellindex_theta = int(match4.group(2))
#                 data_proc_timestep["tracers"].append((local_id,global_id,pos_r,cellindex_r,pos_theta,cellindex_theta))
#                 tracer_data[global_id-1].append((timestep,time,pos_r,pos_theta))
#                 time_data[timestep_p].append((global_id,pos_r,pos_theta,time))
#                 continue
#         tracer_num_record.append(len(data_proc_timestep["tracers"]))
#         data_proc.append(copy.deepcopy(data_proc_timestep))
# #    for ele in data_proc:
# #        wt.write(str(ele)+"\n")
#     data.append(data_proc)
# test_tracer = 900
#
# tracer_data_i = sorted(tracer_data[test_tracer],key=lambda x:x[0])
# T = [x[1] for x in tracer_data_i]
# R = [x[2] for x in tracer_data_i]
# THETA = [x[3] for x in tracer_data_i]
# plt.plot(T,THETA)
# plt.plot(time_steps,tracer_tpos)
# plt.show()
# plt.plot(T,R)
# plt.plot(time_steps,tracer_rpos)
# plt.show()



