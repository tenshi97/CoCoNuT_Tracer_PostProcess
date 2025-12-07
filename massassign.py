import numpy as np

import h5reader
import matplotlib.pyplot as plt
import math
import re
import copy
from collections import deque
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
import diag_2d
queue = deque()
pending_queue = deque()
model_name = "21m_old"
h=h5reader.hydrof(model=model_name,subdirectory="models")
g = open(f"{model_name}_mass_assign.txt","w")

wb = Workbook()
ws = wb.active

xif = h.xzr()
yif = h.yzr()
def grid_finder(x,y):
    st = 0
    ed = r_len - 1
    x_cell = -1
    y_cell = -1
    while(st<ed):
        mid = (st+ed)//2
        if(x<xif[mid]):
            ed = mid
        else:
            st = mid+1
    x_cell = st
    st = 0
    ed = theta_len - 1
    while(st<ed):
        mid = (st+ed)//2
        if(y<yif[mid]):
            ed = mid
        else:
            st = mid+1
    y_cell = st
    return (x_cell,y_cell)

t = h.ngroups-1

h.set_index(t)
#for t in range(0,y.ngroups):
#    y.set_index(t)
#    tracer_ids = np.array([int(round(x)) for x in y.tracer_id()])
#    for i in range(len(tracer_ids)):
#        gid = tracer_ids[i]
#        tracer_data[gid].append((y.time(),y.tracer_x()[i],y.tracer_y()[i]))
r_len = len(h.xzn())
theta_len = len(h.yzn())
empty_dict = {}
array_2D = [[] for x in range(r_len)]
for row in array_2D:
    for i in range(theta_len):
        row.append({})

#Init
tracer_ids = np.array([int(round(x)) for x in h.tracer_id()])
tracer_num = len(tracer_ids)
tracer_data = []
for i in range(tracer_num):
    (x_cell,y_cell) = grid_finder(h.tracer_x()[i],h.tracer_y()[i])
    tracer_data.append((tracer_ids[i],h.tracer_x()[i],h.tracer_y()[i],x_cell,y_cell))
ebind = diag_2d.get_binding_energy(h)
count = 0
for i in range(tracer_num):
    tracer = tracer_data[i]
    id = tracer[0]
    x = tracer[3]
    y = tracer[4]
    if(ebind[x][y][0]<=0):
        print(f"tracer {i} is bounded, skipping")
        continue
    else:
        count +=1
    grid = array_2D[x][y]
    grid[id] = 0
    num = len(grid)
    for key in grid.keys():
        grid[key] = 1/num
print(f"{count} tracers counted")
for x in range(r_len):
    for y in range(theta_len):
        if(len(array_2D[x][y])!=0):
            cell = ws.cell(row=x + 1, column=y + 1)
            color = f"{153:02X}{217:02X}{234:02X}"
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.fill = fill
            cv = ""
            grid = array_2D[x][y]
            for key in grid:
                cv += f"{int(key)}:{grid[key]:.3f}\n"
            cell.value = cv
            queue.append((x,y))
#Init
def get_neighbour_grid(x,y):
    result = []
    if(x-1>=0):
        result.append((x-1,y))
    if(y-1>=0):
        result.append((x,y-1))
    if(x+1<r_len):
        result.append((x+1,y))
    if(y+1<theta_len):
        result.append((x,y+1))
    return result
step = 0
while(True):
    if(len(queue)==0):
        if(len(pending_queue)==0):
            break
        else:
            for grid in pending_queue:
                x = grid[0]
                y = grid[1]
                total = 0
                for value in array_2D[x][y].values():
                    total += value
                for key in array_2D[x][y].keys():
                    array_2D[x][y][key]/=total
            queue.extend(pending_queue)
            #f.write(f"{queue}\n")
            pending_queue.clear()
    step += 1
    grid_pop = queue.popleft()
    xp = grid_pop[0]
    yp = grid_pop[1]
    neighbours = get_neighbour_grid(grid_pop[0],grid_pop[1])
    for grid in neighbours:
        x = grid[0]
        y = grid[1]
        if(len(array_2D[x][y])!=0):
            if(grid in pending_queue):
                array_2D[x][y].update(array_2D[xp][yp])
        else:
            if(not grid in pending_queue):
                pending_queue.append(grid)
                array_2D[x][y].update(array_2D[xp][yp])
Mass_Assignment = [0.0 for x in range(tracer_num)]
c = 29979245800
M_ALL = 0.0
debug_mass = 0.0
for x in range(r_len):
    for y in range(theta_len):
        grid = array_2D[x][y]
        cell = ws.cell(row=x+1,column=y+1)
        cv = ""
        for key in grid:
            cv+=f"{int(key)}:{grid[key]:.3f}\n"
        if cell.value is None or cell.value == "":
            cell.value = cv
        cell.alignment = Alignment(wrapText=True)
        r_1 = h.xzl()[x]
        r_2 = h.xzr()[x]
        r_mid = h.xzn()[x]
        theta_1 = h.yzl()[y]
        theta_2 = h.yzr()[y]
        theta_mid = h.yzn()[y]
        DV = 2*np.pi*(np.cos(theta_1)-np.cos(theta_2))*(r_2**3-r_1**3)/3
        LF = 1.0/np.sqrt(1-(h.vex()[x][y][0]**2+h.vey()[x][y][0]**2)/(c**2))    #Lorentz Factor
        DM = h.den()[x][y][0] * DV * (h.phi()[x][y][0]**6) * LF
        M_ALL+=DM
        if(ebind[x][y][0]<=0):
            print(f"Grid [{x}] [{y}] is bounded,skipping")
            continue
        for key in grid:
            t_id = int(key)
            t_weight = grid[key]
            Mass_Assignment[t_id-1] += t_weight*DM
wb.save("output.xlsx")
Mass_Sun = 1.98855e33   # UNIT:grams
Total_Mass = 0.0
print(M_ALL/Mass_Sun)
for i in range(tracer_num):
    g.write(f"{i+1}: {Mass_Assignment[i]}\n")
    Total_Mass += Mass_Assignment[i]
MS = Total_Mass/Mass_Sun
tracer_alltime_data = [[] for x in range(tracer_num)]
g.write(f"Total Ejected Mass:{Total_Mass}({MS:.3f}of Solar Mass), Total Grid Mass:{M_ALL/Mass_Sun}")



